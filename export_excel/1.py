import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import openai
import os
from datetime import datetime
import time
import logging
from tenacity import retry, stop_after_attempt, wait_exponential

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# OpenAI配置
openai.api_base = "https://text.pollinations.ai/openai/"  # 替换为您的第三方OpenAI API地址
openai.api_key = "sk-*"  # 替换为您的API密钥

# 需要填充的列名
TARGET_COLUMNS = ['数据移动类型', '数据组', '数据属性', 'CFP']
# 功能需求列名
REQUIREMENT_COLUMN = '功能用户需求'
# 批量处理的行数
BATCH_SIZE = 2  # 每次处理2行数据
REFERENCE_START_ROW = 2  # 参考数据开始行
REFERENCE_END_ROW = 20   # 参考数据结束行
PROCESS_START_ROW = 21   # 处理数据开始行

# 数据移动类型的规则说明
DATA_MOVEMENT_RULES = """
数据移动类型的判断规则：
1. 当功能涉及数据输入、新增、录入时，通常是 Entry (E)
2. 当功能涉及数据读取、查询、显示时，通常是 Read (R)
3. 当功能涉及数据写入、更新、修改时，通常是 Write (W)
4. 当功能涉及数据删除时，通常是 Delete (D)
5. 当功能涉及数据导出、下载时，通常是 Exit (X)

数据组和数据属性的判断规则：
1. 数据组通常是功能操作的主要对象（如：用户、角色、配置等）
2. 数据属性是该数据组的具体属性（如：ID、名称、状态等）
3. 根据功能描述来确定涉及的数据对象和属性

CFP的计算规则：
1. 每个基本功能点计数为1个CFP
2. 复杂的数据移动（涉及多个数据组）可能会有更高的CFP值
"""


def setup_logging():
    """设置日志"""
    log_dir = "logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    log_file = os.path.join(log_dir, f"excel_process_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    file_handler = logging.FileHandler(log_file)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logging.getLogger().addHandler(file_handler)


def get_cell_value(cell):
    """安全地获取单元格值"""
    try:
        return cell.value
    except:
        return None


def is_merged_cell(cell):
    """检查是否是合并单元格"""
    try:
        return isinstance(cell, openpyxl.cell.cell.MergedCell)
    except:
        return False


def get_merged_cell_value(sheet, row, col):
    """获取合并单元格的值"""
    try:
        for merged_range in sheet.merged_cells.ranges:
            if (row, col) in merged_range:
                # 获取合并区域的主单元格（左上角）
                main_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return get_cell_value(main_cell)
    except:
        pass
    return None


def get_column_index(sheet, column_name):
    """获取列名对应的列索引"""
    for col in range(1, sheet.max_column + 1):
        cell_value = get_cell_value(sheet.cell(row=1, column=col))
        if cell_value == column_name:
            return col
    return None


def get_column_name(sheet, col_idx):
    """获取列索引对应的列名"""
    return get_cell_value(sheet.cell(row=1, column=col_idx))


def read_excel(file_path):
    """读取Excel文件"""
    try:
        # 读取Excel文件，保持原始格式
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        logging.info(f"成功打开Excel文件: {file_path}")
        logging.info(f"工作表名称: {sheet.title}")
        logging.info(f"总行数: {sheet.max_row}, 总列数: {sheet.max_column}")

        # 获取需要填充的单元格信息
        empty_cells = []
        reference_data = []

        # 获取目标列的索引
        target_cols = []
        for col_name in TARGET_COLUMNS:
            col_idx = get_column_index(sheet, col_name)
            if col_idx:
                target_cols.append(col_idx)
                logging.info(f"找到列 '{col_name}' 在第 {col_idx} 列")
            else:
                logging.error(f"未找到列: {col_name}")
                raise ValueError(f"未找到列: {col_name}")

        # 获取功能需求列索引
        req_col_idx = get_column_index(sheet, REQUIREMENT_COLUMN)
        if not req_col_idx:
            logging.error(f"未找到列: {REQUIREMENT_COLUMN}")
            raise ValueError(f"未找到列: {REQUIREMENT_COLUMN}")
        logging.info(f"找到功能需求列在第 {req_col_idx} 列")

        # 获取前20行作为参考数据
        reference_count = 0
        for row in range(REFERENCE_START_ROW, REFERENCE_END_ROW + 1):
            row_data = []
            has_value = False
            for col in target_cols:
                cell = sheet.cell(row=row, column=col)
                if is_merged_cell(cell):
                    cell_value = get_merged_cell_value(sheet, row, col)
                else:
                    cell_value = get_cell_value(cell)
                if cell_value and str(cell_value).strip():
                    has_value = True
                row_data.append(cell_value)
            
            if has_value:
                # 获取功能需求
                req_cell = sheet.cell(row=row, column=req_col_idx)
                if is_merged_cell(req_cell):
                    req_value = get_merged_cell_value(sheet, row, req_col_idx)
                else:
                    req_value = get_cell_value(req_cell)
                
                if req_value:
                    reference_data.append({
                        'requirement': req_value,
                        'values': row_data
                    })
                    reference_count += 1
                    logging.info(f"收集到第 {row} 行的参考数据，功能需求：{req_value}")

        logging.info(f"共收集到 {reference_count} 条参考数据")

        # 检查需要处理的行（从第21行开始）
        empty_count = 0
        for row in range(PROCESS_START_ROW, sheet.max_row + 1):
            row_has_empty = False
            empty_cols = []
            
            # 检查目标列是否有空值
            for col in target_cols:
                cell = sheet.cell(row=row, column=col)
                if is_merged_cell(cell):
                    cell_value = get_merged_cell_value(sheet, row, col)
                else:
                    cell_value = get_cell_value(cell)
                
                if cell_value is None or str(cell_value).strip() == '':
                    row_has_empty = True
                    empty_cols.append(col)
                    logging.info(f"第 {row} 行第 {col} 列需要填充")
            
            if row_has_empty:
                # 获取功能需求
                req_cell = sheet.cell(row=row, column=req_col_idx)
                if is_merged_cell(req_cell):
                    req_value = get_merged_cell_value(sheet, row, req_col_idx)
                else:
                    req_value = get_cell_value(req_cell)
                
                if req_value:
                    empty_cells.append({
                        'row': row,
                        'empty_columns': empty_cols,
                        'requirement': req_value
                    })
                    empty_count += 1
                    logging.info(f"找到需要填充的行：{row}，功能需求：{req_value}")

        logging.info(f"共找到 {empty_count} 行需要填充")
        return workbook, sheet, empty_cells, reference_data
    except Exception as e:
        logging.error(f"读取Excel文件失败: {str(e)}")
        raise


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def generate_content(prompt, reference_data):
    """使用OpenAI API生成内容，添加重试机制"""
    try:
        # 构建参考数据文本
        reference_text = "\n示例数据：\n"
        for ref in reference_data[:5]:  # 只使用前5个有效的参考数据
            values_str = ", ".join(str(x) for x in ref['values'] if x)
            reference_text += f"功能需求: {ref['requirement']}\n对应的值: {values_str}\n"

        # 添加规则说明
        full_prompt = DATA_MOVEMENT_RULES + "\n\n" + reference_text + "\n" + prompt

        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system",
                 "content": """你是一个专业的功能点分析专家，请根据功能需求描述和规则来生成数据移动类型、数据组、数据属性、CFP的值。
注意：
1. 严格按照数据移动类型的规则判断（E/R/W/X/D）
2. 数据组和数据属性要与功能描述密切相关
3. CFP值要反映功能的复杂度
4. 保持数据的一致性和准确性
5. 每行数据必须包含这4个值，用逗号分隔
6. 每行数据单独一行输出"""},
                {"role": "user", "content": full_prompt}
            ],
            temperature=0.2,  # 降低温度以提高一致性
            max_tokens=1000
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logging.error(f"调用OpenAI API失败: {str(e)}")
        raise


def process_excel(input_file, output_file):
    """处理Excel文件"""
    try:
        logging.info("开始处理Excel文件")

        # 读取Excel
        workbook, sheet, empty_cells, reference_data = read_excel(input_file)
        
        if not empty_cells:
            logging.warning("没有找到需要填充的单元格")
            workbook.save(output_file)
            return
            
        if not reference_data:
            logging.warning("没有找到参考数据")
            workbook.save(output_file)
            return

        # 设置绿色填充
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        # 批量处理数据
        total_batches = (len(empty_cells) + BATCH_SIZE - 1) // BATCH_SIZE
        current_batch = 0
        
        for i in range(0, len(empty_cells), BATCH_SIZE):
            current_batch += 1
            batch = empty_cells[i:i + BATCH_SIZE]
            logging.info(f"处理第 {current_batch}/{total_batches} 批数据，包含 {len(batch)} 行")
            
            try:
                # 构建批量提示词
                batch_data = []
                for cell_info in batch:
                    batch_data.append(f"行 {cell_info['row']} 的功能需求：\n{cell_info['requirement']}")

                prompt = f"请根据以下功能需求，生成对应的数据移动类型、数据组、数据属性、CFP：\n\n" + "\n\n".join(batch_data)
                logging.info(f"发送批次 {current_batch} 到API进行处理")

                generated_content = generate_content(prompt, reference_data)
                
                if generated_content:
                    logging.info(f"API返回内容：\n{generated_content}")
                    # 处理每行生成的内容
                    content_lines = generated_content.split('\n')
                    for j, cell_info in enumerate(batch):
                        if j < len(content_lines):
                            # 记录生成的内容
                            logging.info(f"行 {cell_info['row']} 的功能需求：{cell_info['requirement']}")
                            logging.info(f"生成的内容：{content_lines[j].strip()}")
                            
                            content_parts = content_lines[j].strip().split(',')
                            for k, col in enumerate(cell_info['empty_columns']):
                                if k < len(content_parts):
                                    try:
                                        target_cell = sheet.cell(row=cell_info['row'], column=col)
                                        if not is_merged_cell(target_cell):  # 只填充非合并单元格
                                            target_cell.value = content_parts[k].strip()
                                            target_cell.fill = green_fill
                                            logging.info(f"成功填充单元格: 行 {cell_info['row']}, 列 {col}, 值: {content_parts[k].strip()}")
                                    except Exception as cell_error:
                                        logging.warning(f"无法填充单元格: 行 {cell_info['row']}, 列 {col}, 错误: {str(cell_error)}")

                # 避免API限制
                time.sleep(1)
            except Exception as e:
                logging.error(f"处理批次 {current_batch} 失败: 行 {batch[0]['row']} 到 {batch[-1]['row']}, 错误: {str(e)}")
                continue

        # 保存结果
        workbook.save(output_file)
        logging.info(f"Excel处理完成，已保存到: {output_file}")

    except Exception as e:
        logging.error(f"处理Excel文件失败: {str(e)}")
        raise


def main():
    """主函数"""
    setup_logging()

    input_file = "附件4：金科研发云测试平台二期项目-功能拆分表20250318.xlsx"  # 替换为您的输入文件路径
    output_file = f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        process_excel(input_file, output_file)
        print(f"处理完成！输出文件: {output_file}")
    except Exception as e:
        print(f"处理失败: {str(e)}")


if __name__ == "__main__":
    main()